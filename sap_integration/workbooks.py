from openpyxl import load_workbook


def same_nonempty_workbook_content(first_path, second_path):
    """Return whether two SAP exports contain the same non-empty cell data.

    XLSX files can have different ZIP metadata even when their worksheets are
    identical, so comparing file hashes is not sufficient here.
    """
    first_rows = _workbook_rows(first_path)
    second_rows = _workbook_rows(second_path)
    return first_rows == second_rows and any(
        any(value not in (None, "") for value in row)
        for row in first_rows[1:]
    )


def _workbook_rows(path):
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        return tuple(
            tuple(cell for cell in row)
            for row in workbook.active.iter_rows(values_only=True)
        )
    finally:
        workbook.close()
